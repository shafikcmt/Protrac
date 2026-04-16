from __future__ import annotations

from datetime import date
from typing import List, Optional

from django.http import HttpResponse
from django.utils.dateparse import parse_date

from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from tracking.models import ProductionLine, LineTarget

# import the function from your existing slide-3 logic file
from common.utils.time import today


class SewingLineDashboardV2QualityExportView(APIView):
    """
    Slide-3 Export: Hourly Quality Monitoring
    Exports: hour, dhu, defects, remarks, defect_breakdown
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from tracking.api.sewing_dashboard_v2 import _get_hourly_quality_rows
        # ----------------------------
        # Parse params
        # ----------------------------
        d = parse_date(request.query_params.get("date") or "") or today()

        # accept multiple param styles
        line_ids: List[str] = (
            request.query_params.getlist("production_line_ids[]")
            or request.query_params.getlist("production_line_ids")
            or request.query_params.getlist("line_ids[]")
            or request.query_params.getlist("line_ids")
        )

        # optional filters (same as API)
        order_id = request.query_params.get("order_id")
        style_id = request.query_params.get("style_id")
        buyer_id = request.query_params.get("buyer_id")

        # sizes/colors often come as arrays; slide-3 function expects single
        size = (request.query_params.getlist("sizes[]") or request.query_params.getlist("sizes") or [None])[0]
        color = (request.query_params.getlist("colors[]") or request.query_params.getlist("colors") or [None])[0]

        # convert to int safely
        def to_int(x):
            try:
                return int(x)
            except Exception:
                return None

        order_id = to_int(order_id)
        style_id = to_int(style_id)
        buyer_id = to_int(buyer_id)

        # ----------------------------
        # Resolve lines
        # ----------------------------
        qs = ProductionLine.objects.all().order_by("name")

        if line_ids:
            ids_int = [to_int(x) for x in line_ids]
            ids_int = [x for x in ids_int if x is not None]
            qs = qs.filter(id__in=ids_int)

        # ✅ important: if no line selected, export ALL (avoid 404 like before)
        lines = list(qs)
        if not lines:
            return Response({"detail": "No production lines found for export."}, status=status.HTTP_400_BAD_REQUEST)

        # ----------------------------
        # Build workbook
        # ----------------------------
        wb = Workbook()
        ws = wb.active
        ws.title = "Hourly Quality"

        headers = [
            "Date",
            "Line",
            "Hour",
            "Units Checked",
            "Defects",
            "DHU %",
            "Remarks",
            "Defect Breakdown",
        ]
        ws.append(headers)

        # column widths (readable)
        widths = [12, 18, 6, 14, 10, 10, 20, 50]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ----------------------------
        # Fill data
        # ----------------------------
        for line in lines:
            line_target = LineTarget.objects.filter(production_line=line, target_date=d).first()

            rows = _get_hourly_quality_rows(
                production_line=line,
                target_date=d,
                line_target=line_target,
                order_id=order_id,
                style_id=style_id,
                buyer_id=buyer_id,
                size=size,
                color=color,
            )

            for r in rows:
                remarks = ", ".join(r.get("remarks") or [])
                # defect_breakdown => "OIL SPOT(4); SKIP STC(1)"
                breakdown_items = []
                for it in (r.get("defect_breakdown") or []):
                    nm = it.get("name") or it.get("code") or ""
                    qty = it.get("qty") or 0
                    breakdown_items.append(f"{nm}({qty})")
                breakdown = "; ".join(breakdown_items)

                ws.append([
                    d.isoformat(),
                    getattr(line, "name", f"Line {line.id}"),
                    r.get("hour"),
                    r.get("units", 0),
                    r.get("defects", 0),
                    float(r.get("dhu") or 0.0),
                    remarks,
                    breakdown,
                ])

        # ----------------------------
        # Return as xlsx
        # ----------------------------
        filename = f"sewing_quality_slide3_{d.isoformat()}.xlsx"

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp