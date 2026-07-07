"""
Tests for the automated Daily Production Report email.

The heavy report-data query is patched with a small fixed fixture so these tests
focus on the mail behaviour: To/Cc splitting, inactive exclusion, the .xlsx
attachment, and the "no active to-recipient → skip" rule. Email is captured with
Django's locmem backend.
"""

from datetime import date
from io import BytesIO
from unittest.mock import patch

import pytest
from django.core import mail
from openpyxl import load_workbook

from tracking.models import MailRecipient
from tracking.models.mail import RecipientType, ReportType

MAIL_MODULE = "tracking.services.report.daily_production_mail"

REPORT_DATE = date(2026, 7, 7)

# One line, two buyers so the buyer breakdown / grouping paths are exercised.
FAKE_REPORT_DATA = [
    {
        "production_line_id": 1,
        "production_line_name": "Sewing-1",
        "orders": [
            {
                "line": "Sewing-1", "buyer": "BATA", "style": "ST-100",
                "order_quantity": 100, "working_hours": 8.0, "working_days": 1,
                "input": 80,
                "front": {"day": 20, "cumulative": 80},
                "back": {"day": 20, "cumulative": 80},
                "sleeve": {"day": 20, "cumulative": 80},
                "hood": {"day": 5, "cumulative": 20},
                "collar": {"day": 5, "cumulative": 20},
                "lining": {"day": 10, "cumulative": 40},
                "assembly_input": {"day": 15, "cumulative": 60},
                "output": {"day": 30, "cumulative": 60},
                "inspection": {"day": 32, "cumulative": 62},
                "packed": {"day": 25, "cumulative": 50},
                "dhu_day": 1.5, "dhu_average": 1.2,
                "is_hidden": False, "is_pending_transition": False, "remarks": "",
            },
            {
                "line": "Sewing-1", "buyer": "HUGO BOSS", "style": "ST-200",
                "order_quantity": 50, "working_hours": None, "working_days": 1,
                "input": 40,
                "front": {"day": 10, "cumulative": 40},
                "back": {"day": 10, "cumulative": 40},
                "sleeve": {"day": 10, "cumulative": 40},
                "hood": {"day": 0, "cumulative": 0},
                "collar": {"day": 0, "cumulative": 0},
                "lining": {"day": 0, "cumulative": 0},
                "assembly_input": {"day": 8, "cumulative": 30},
                "output": {"day": 12, "cumulative": 24},
                "inspection": {"day": 13, "cumulative": 25},
                "packed": {"day": 10, "cumulative": 20},
                "dhu_day": 0.0, "dhu_average": 0.0,
                "is_hidden": False, "is_pending_transition": False, "remarks": "",
            },
        ],
    }
]


@pytest.fixture
def patched_report_data():
    with patch(
        f"{MAIL_MODULE}.get_daily_production_report_data",
        return_value=FAKE_REPORT_DATA,
    ):
        yield


def _make_recipient(email, recipient_type=RecipientType.TO, is_active=True,
                    report_type=ReportType.DAILY_PRODUCTION):
    return MailRecipient.objects.create(
        email=email, recipient_type=recipient_type, is_active=is_active,
        report_type=report_type,
    )


@pytest.mark.django_db
class TestDailyProductionMail:
    @pytest.fixture(autouse=True)
    def _mail_settings(self, settings):
        # locmem backend captures email in mail.outbox regardless of DEBUG.
        settings.EMAIL_BACKEND = "django.core.mail.backends.locmem.EmailBackend"
        settings.DEFAULT_FROM_EMAIL = "reports@example.com"

    def _send(self):
        from tracking.services.report.daily_production_mail import (
            send_daily_production_report_email,
        )
        return send_daily_production_report_email(report_date=REPORT_DATE)

    def test_to_and_cc_split(self, patched_report_data):
        _make_recipient("manager@example.com", RecipientType.TO)
        _make_recipient("lead@example.com", RecipientType.TO)
        _make_recipient("director@example.com", RecipientType.CC)

        sent = self._send()

        assert sent is True
        assert len(mail.outbox) == 1
        msg = mail.outbox[0]
        assert set(msg.to) == {"manager@example.com", "lead@example.com"}
        assert set(msg.cc) == {"director@example.com"}

    def test_inactive_recipients_excluded(self, patched_report_data):
        _make_recipient("active@example.com", RecipientType.TO)
        _make_recipient("gone@example.com", RecipientType.TO, is_active=False)
        _make_recipient("oldcc@example.com", RecipientType.CC, is_active=False)

        self._send()

        msg = mail.outbox[0]
        assert msg.to == ["active@example.com"]
        assert msg.cc == []

    def test_xlsx_attachment_present_and_valid(self, patched_report_data):
        _make_recipient("manager@example.com", RecipientType.TO)

        self._send()

        msg = mail.outbox[0]
        assert len(msg.attachments) == 1
        filename, content, mimetype = msg.attachments[0]
        assert filename == f"Daily_Production_Report_{REPORT_DATE}.xlsx"
        assert mimetype == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # The attachment is a real, openable workbook.
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        assert ws["A1"].value == "Humana Apparels Pvt. Ltd."
        # Two-tier header includes the Lining group (col 16 on the group row).
        assert ws.cell(row=6, column=16).value == "Lining"

    def test_html_and_text_alternatives(self, patched_report_data):
        _make_recipient("manager@example.com", RecipientType.TO)

        self._send()

        msg = mail.outbox[0]
        # HTML alternative attached; plain-text body present as fallback.
        assert any(ct == "text/html" for _, ct in msg.alternatives)
        assert "Daily Production Report" in msg.body
        assert "BATA" in msg.body  # per-buyer breakdown rendered

    def test_no_active_to_recipient_skips_send(self, patched_report_data):
        # Only a CC recipient — no active "to" means no send.
        _make_recipient("director@example.com", RecipientType.CC)

        sent = self._send()

        assert sent is False
        assert len(mail.outbox) == 0
