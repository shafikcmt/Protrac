"""
Automated Daily Production Report email: server-side Excel generation and send.

The Excel layout deliberately mirrors the frontend "Export Excel" output in
frontend/lib/reports/daily-production-report.ts — title block, two-tier
Day/Cumm group headers (including the Lining column), buyer subtotal rows, and a
grand-total footer — so the emailed file looks identical to what a user
downloads from the report page. All numbers are pre-computed from the same
report-data service the API/frontend use; no Excel formulas are written, so
#REF!/#DIV/0! error classes cannot occur.
"""

from __future__ import annotations

import logging
import smtplib
import time
from datetime import date
from email.mime.image import MIMEImage
from io import BytesIO
from typing import Any, Dict, List, Optional

from django.conf import settings
from django.contrib.staticfiles import finders
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common.utils.time import today
from tracking.models import MailRecipient
from tracking.models.mail import RecipientType, ReportType
from tracking.services.report import get_daily_production_report_data

logger = logging.getLogger("tracking.reports.daily_production_mail")

COMPANY_NAME = "Humana Apparels Pvt. Ltd."

# Horizontal wordmark, resolved from the tracking app's static/ dir. Optional:
# if the file isn't present the Excel/email simply render without a logo, so a
# missing asset never breaks report delivery.
LOGO_STATIC_PATH = "images/company-logo-horizontal.png"
LOGO_WIDTH_PX = 160
LOGO_HEIGHT_PX = 73  # keeps the 220×101 source aspect ratio


def _logo_path() -> Optional[str]:
    """Absolute filesystem path to the logo, or None if it isn't installed."""
    try:
        return finders.find(LOGO_STATIC_PATH)
    except Exception:  # noqa: BLE001 — a logo lookup must never break a send
        return None

# ── Shared styling (mirrors the frontend export's blue header + grey subtotal) ──
HEADER_FILL = PatternFill("solid", fgColor="2563EB")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E5E7EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
_THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(top=_THIN, left=_THIN, bottom=_THIN, right=_THIN)
PCT_FMT = '0.00"%"'
EM_DASH = "–"

LAST_COL = 30  # 7 single + 11 two-column groups + Remarks


def _num(v: Any) -> float:
    try:
        n = float(v)
    except (TypeError, ValueError):
        return 0
    return n if n == n else 0  # guard against NaN


def _efficiency(output_part: float, input_val: float) -> float:
    return (output_part / input_val * 100) if input_val > 0 else 0


# ---------------------------------------------------------------------------
# Buyer grouping — ported verbatim from the frontend production-report-table so
# the exported rows/subtotals match the on-screen table one-for-one. Grouping is
# purely by the buyer string, so a distinct buyer (e.g. "Sample") automatically
# gets its own group + subtotal.
# ---------------------------------------------------------------------------
_METRICS = [
    "front",
    "back",
    "sleeve",
    "hood",
    "collar",
    "lining",
    "assembly_input",
    "output",
    "inspection",
    "packed",
]


def build_display_by_buyer(report_data: List[Dict[str, Any]]) -> Dict[str, List[dict]]:
    orders_by_buyer: Dict[str, List[dict]] = {}
    hidden_by_buyer: Dict[str, List[dict]] = {}

    for line in report_data:
        for order in line.get("orders", []):
            buyer = order.get("buyer") or "UNKNOWN"
            style = order.get("style") or "UNKNOWN"
            line_name = order.get("line") or line.get("production_line_name") or "UNKNOWN"
            is_hidden = bool(order.get("is_hidden"))
            target = hidden_by_buyer if is_hidden else orders_by_buyer
            bucket = target.setdefault(buyer, [])

            key = f"{line_name}||{style}"
            acc = next((x for x in bucket if x["__key"] == key), None)
            if acc is None:
                acc = {
                    "__key": key,
                    "line": line_name,
                    "buyer": buyer,
                    "style": style,
                    "is_hidden": is_hidden,
                    "order_quantity": 0,
                    "input": 0,
                    "working_days": order.get("working_days") or 1,
                    "working_hours": order.get("working_hours"),
                    "__dhu_day_num": 0,
                    "__dhu_day_den": 0,
                    "__dhu_avg_num": 0,
                    "__dhu_avg_den": 0,
                    "dhu_day": 0,
                    "dhu_average": 0,
                }
                for m in _METRICS:
                    acc[m] = {"day": 0, "cumulative": 0}
                bucket.append(acc)

            acc["order_quantity"] += _num(order.get("order_quantity"))
            acc["input"] += _num(order.get("input"))
            for m in _METRICS:
                src = order.get(m) or {}
                acc[m]["day"] += _num(src.get("day"))
                acc[m]["cumulative"] += _num(src.get("cumulative"))

            out_day = _num((order.get("output") or {}).get("day"))
            out_cum = _num((order.get("output") or {}).get("cumulative"))
            if out_day > 0:
                acc["__dhu_day_num"] += _num(order.get("dhu_day")) * out_day
                acc["__dhu_day_den"] += out_day
            if out_cum > 0:
                acc["__dhu_avg_num"] += _num(order.get("dhu_average")) * out_cum
                acc["__dhu_avg_den"] += out_cum
            acc["dhu_day"] = (
                acc["__dhu_day_num"] / acc["__dhu_day_den"]
                if acc["__dhu_day_den"] > 0
                else 0
            )
            acc["dhu_average"] = (
                acc["__dhu_avg_num"] / acc["__dhu_avg_den"]
                if acc["__dhu_avg_den"] > 0
                else 0
            )

    display: Dict[str, List[dict]] = {}
    for b in list(orders_by_buyer.keys()) + [
        k for k in hidden_by_buyer if k not in orders_by_buyer
    ]:
        display[b] = orders_by_buyer.get(b, []) + hidden_by_buyer.get(b, [])
    return display


def _buyer_daily_output(orders: List[dict]) -> int:
    """Sum of visible orders' daily output for the per-buyer email summary."""
    return int(
        sum(o["output"]["day"] for o in orders if not o.get("is_hidden"))
    )


def _buyer_cumulative_output(orders: List[dict]) -> int:
    """Sum of visible orders' cumulative output (running total) per buyer."""
    return int(
        sum(o["output"]["cumulative"] for o in orders if not o.get("is_hidden"))
    )


# ---------------------------------------------------------------------------
# Excel generation (openpyxl) — returns in-memory bytes.
# ---------------------------------------------------------------------------
_COL_WIDTHS = [
    10, 16, 18, 11, 8, 7, 10,  # Line..Input
    9, 9, 9, 9, 9, 9, 11, 11, 9, 9, 11, 11, 10, 10,  # Front..Output
    11, 11, 9, 9, 11, 11, 10, 10,  # Efficiency..Packed
    28,  # Remarks
]

# [group label, left col (1-based), day-sub, right-sub]
_GROUPS = [
    ("Front", 8, "Day", "Cumm"),
    ("Back", 10, "Day", "Cumm"),
    ("Sleeve", 12, "Day", "Cumm"),
    ("Hood/Collar", 14, "Day", "Cumm"),
    ("Lining", 16, "Day", "Cumm"),
    ("Assembly Input", 18, "Day", "Cumm"),
    ("Output", 20, "Day", "Cumm"),
    ("Efficiency %", 22, "Day", "Avg"),
    ("DHU %", 24, "Day", "Avg"),
    ("Inspection", 26, "Day", "Cumm"),
    ("Packed", 28, "Day", "Cumm"),
]
_SINGLE_COLS = [
    (1, "Line"),
    (2, "Buyer"),
    (3, "Style"),
    (4, "Order Qty"),
    (5, "W Days"),
    (6, "Hrs"),
    (7, "Input"),
    (30, "Remarks"),
]


def _style_header_cell(cell, horizontal="center"):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def build_daily_production_xlsx(
    report_data: List[Dict[str, Any]],
    summary: Dict[str, Any],
    report_date: date,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Production"

    for i, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ── Title block (rows 1-4) + spacer (row 5) ──
    # Optional top-left logo floats over rows 1-4; when present, the text block
    # shifts right (col 4) so it sits beside the logo instead of behind it. No
    # rows are added, so freeze panes / row indices below stay put.
    title_col = 1
    logo_path = _logo_path()
    if logo_path:
        try:
            img = XLImage(logo_path)
            img.width, img.height = LOGO_WIDTH_PX, LOGO_HEIGHT_PX
            ws.add_image(img, "A1")
            title_col = 4
        except Exception:  # noqa: BLE001 — a bad image must not break the file
            logger.warning("Could not embed logo into Excel from %s", logo_path)
            title_col = 1

    ws.cell(row=1, column=title_col, value=COMPANY_NAME).font = Font(bold=True, size=14)
    ws.cell(row=2, column=title_col, value="Daily Production Report").font = Font(bold=True, size=12)
    ws.cell(row=3, column=title_col, value=f"Report Date: {report_date}").font = Font(size=10, color="374151")
    ws.cell(row=4, column=title_col, value="Generated automatically by Production Tracking Software").font = Font(size=10, color="6B7280")

    # ── Two-tier header (rows 6-7) ──
    group_row, sub_row = 6, 7
    for col, label in _SINGLE_COLS:
        ws.merge_cells(start_row=group_row, start_column=col, end_row=sub_row, end_column=col)
        ws.cell(row=group_row, column=col, value=label)
    for label, left, sub_l, sub_r in _GROUPS:
        ws.merge_cells(start_row=group_row, start_column=left, end_row=group_row, end_column=left + 1)
        ws.cell(row=group_row, column=left, value=label)
        ws.cell(row=sub_row, column=left, value=sub_l)
        ws.cell(row=sub_row, column=left + 1, value=sub_r)
    for r in (group_row, sub_row):
        for c in range(1, LAST_COL + 1):
            _style_header_cell(ws.cell(row=r, column=c))

    ws.freeze_panes = "A8"  # keep title + headers visible on scroll

    # ── Data rows grouped by buyer, each followed by a subtotal ──
    display = build_display_by_buyer(report_data)
    row_idx = 8

    def write_data_row(order: dict):
        nonlocal row_idx
        input_val = _num(order.get("input"))
        hood_day = order["hood"]["day"] + order["collar"]["day"]
        hood_cum = order["hood"]["cumulative"] + order["collar"]["cumulative"]
        wh = order.get("working_hours")
        values = [
            order.get("line"),
            order.get("buyer"),
            order.get("style"),
            int(_num(order.get("order_quantity"))),
            order.get("working_days"),
            EM_DASH if wh is None else round(float(wh), 2),
            int(input_val),
            order["front"]["day"], order["front"]["cumulative"],
            order["back"]["day"], order["back"]["cumulative"],
            order["sleeve"]["day"], order["sleeve"]["cumulative"],
            hood_day, hood_cum,
            order["lining"]["day"], order["lining"]["cumulative"],
            order["assembly_input"]["day"], order["assembly_input"]["cumulative"],
            order["output"]["day"], order["output"]["cumulative"],
            _efficiency(order["output"]["day"], input_val),
            _efficiency(order["output"]["cumulative"], input_val),
            _num(order.get("dhu_day")), _num(order.get("dhu_average")),
            order["inspection"]["day"], order["inspection"]["cumulative"],
            order["packed"]["day"], order["packed"]["cumulative"],
            _remarks_text(order),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.border = BORDER
            if 4 <= c <= 29:
                cell.alignment = Alignment(horizontal="right")
            if 22 <= c <= 25:
                cell.number_format = PCT_FMT
            if order.get("is_hidden"):
                cell.font = Font(color="9CA3AF")
        row_idx += 1

    def write_subtotal_row(buyer: str, orders: List[dict]):
        nonlocal row_idx
        visible = [o for o in orders if not o.get("is_hidden")]
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        ws.cell(row=row_idx, column=1, value=f"{buyer} Total")
        ws.cell(row=row_idx, column=4, value=int(sum(o["order_quantity"] for o in visible)))
        ws.cell(row=row_idx, column=7, value=int(sum(o["input"] for o in visible)))
        ws.cell(row=row_idx, column=20, value=int(sum(o["output"]["day"] for o in visible)))
        ws.cell(row=row_idx, column=21, value=int(sum(o["output"]["cumulative"] for o in visible)))
        for c in range(1, LAST_COL + 1):
            cell = ws.cell(row=row_idx, column=c)
            cell.font = BOLD_FONT
            cell.border = BORDER
            cell.fill = SUBTOTAL_FILL
            cell.alignment = Alignment(horizontal="right" if 4 <= c <= 29 else "left")
        row_idx += 1

    for buyer, orders in display.items():
        for order in orders:
            write_data_row(order)
        write_subtotal_row(buyer, orders)

    # ── Grand-total footer (same numbers as the on-screen summary strip) ──
    row_idx += 1  # spacer
    footer_labels = [
        "Total Lines", "Total Orders", "Order Qty", "Total Input",
        "Daily Output", "Daily Inspection", "Daily Packed", "Efficiency %",
    ]
    footer_values = [
        summary.get("total_production_lines", 0),
        summary.get("total_orders", 0),
        summary.get("total_order_quantity", 0),
        summary.get("daily_input", 0),
        summary.get("daily_output", 0),
        summary.get("daily_inspection", 0),
        summary.get("daily_packed", 0),
        _num(summary.get("overall_efficiency", 0)),
    ]
    for i, label in enumerate(footer_labels, start=1):
        _style_header_cell(ws.cell(row=row_idx, column=i, value=label))
    # Taller row so wrapped labels (e.g. "Daily Inspection") show cleanly on two
    # lines instead of looking cramped — without widening the shared data columns.
    ws.row_dimensions[row_idx].height = 30
    for i, value in enumerate(footer_values, start=1):
        cell = ws.cell(row=row_idx + 1, column=i, value=value)
        cell.font = BOLD_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")
        if i == 8:
            cell.number_format = PCT_FMT

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _remarks_text(order: dict) -> str:
    if order.get("is_hidden"):
        return "Hidden"
    remarks = order.get("remarks")
    if order.get("is_pending_transition") and remarks:
        return str(remarks).replace(" | ", " · ")
    return ""


# ---------------------------------------------------------------------------
# Send
# ---------------------------------------------------------------------------
# Transient SMTP hiccups (dropped socket, brief network blip) are retried a few
# times before we give up. Only connection-level errors are retried — an auth or
# recipient error would just fail identically on every attempt.
_SEND_MAX_RETRIES = 2  # i.e. up to 3 attempts total
_SEND_RETRY_DELAY_SECONDS = 3
_RETRYABLE_SEND_ERRORS = (smtplib.SMTPServerDisconnected, ConnectionError)


def _send_email_with_retry(email: EmailMultiAlternatives, report_date: date) -> None:
    """
    Send ``email``, retrying only on transient connection-level SMTP failures.

    Each attempt and its outcome is logged. Raises the last error if every
    attempt fails, so the caller's ``except`` block records the final failure.
    """
    attempts = _SEND_MAX_RETRIES + 1
    for attempt in range(1, attempts + 1):
        try:
            email.send()
            if attempt > 1:
                logger.info(
                    "Daily production report email sent for %s on attempt %d/%d.",
                    report_date, attempt, attempts,
                )
            return
        except _RETRYABLE_SEND_ERRORS as exc:
            if attempt < attempts:
                logger.warning(
                    "Daily production report email send attempt %d/%d for %s failed "
                    "(%s: %s); retrying in %ds.",
                    attempt, attempts, report_date,
                    type(exc).__name__, exc, _SEND_RETRY_DELAY_SECONDS,
                )
                time.sleep(_SEND_RETRY_DELAY_SECONDS)
            else:
                logger.error(
                    "Daily production report email send failed for %s after %d "
                    "attempts (%s: %s); giving up.",
                    report_date, attempts, type(exc).__name__, exc,
                )
                raise


def send_daily_production_report_email(report_date: Optional[date] = None) -> bool:
    """
    Build and send today's Daily Production Report (all lines/buyers combined)
    to the active MailRecipient rows for report_type="daily_production".

    Returns True if an email was sent, False if skipped (no active "to"
    recipients) or on error. Never raises — a mail failure must not crash the
    scheduler or a request.
    """
    if report_date is None:
        report_date = today()

    try:
        report_data = get_daily_production_report_data(report_date=report_date)

        # Reuse the API's summary generator so the emailed footer/summary numbers
        # match the report page exactly. Imported locally to avoid a circular
        # import (the API view imports this services package at module load).
        from tracking.api.report.daily_production_report import _generate_summary

        summary = _generate_summary(report_data)

        # Cumulative (running-total) output alongside the summary's daily output.
        # Mirror _generate_summary's rule of excluding hidden rows so the two
        # numbers on the Summary card stay directly comparable.
        summary["total_output_cumulative"] = int(
            sum(
                (order.get("output") or {}).get("cumulative", 0) or 0
                for line in report_data
                for order in line.get("orders", [])
                if not order.get("is_hidden")
            )
        )

        recipients = MailRecipient.objects.filter(
            report_type=ReportType.DAILY_PRODUCTION, is_active=True
        )
        to_list = [r.email for r in recipients if r.recipient_type == RecipientType.TO]
        cc_list = [r.email for r in recipients if r.recipient_type == RecipientType.CC]

        if not to_list:
            logger.warning(
                "Daily production report email skipped for %s: no active 'to' recipients.",
                report_date,
            )
            return False

        display = build_display_by_buyer(report_data)
        buyer_breakdown = [
            {
                "buyer": buyer,
                "daily_output": _buyer_daily_output(orders),
                "cumulative_output": _buyer_cumulative_output(orders),
            }
            for buyer, orders in display.items()
        ]

        logo_path = _logo_path()
        context = {
            "company_name": COMPANY_NAME,
            "report_date": report_date,
            "summary": summary,
            "buyer_breakdown": buyer_breakdown,
            "has_logo": bool(logo_path),
        }
        html_body = render_to_string("emails/daily_production_report.html", context)
        text_body = strip_tags(html_body)

        xlsx_bytes = build_daily_production_xlsx(report_data, summary, report_date)
        filename = f"Daily_Production_Report_{report_date}.xlsx"

        email = EmailMultiAlternatives(
            subject=f"Daily Production Report — {report_date}",
            body=text_body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=to_list,
            cc=cc_list,
        )
        email.attach_alternative(html_body, "text/html")

        # Inline logo via Content-ID (cid:) — the most reliable way to show an
        # image in Outlook/Office 365, which block remote and base64 images by
        # default. "related" nests the image with the HTML part.
        if logo_path:
            try:
                with open(logo_path, "rb") as fh:
                    logo_img = MIMEImage(fh.read())
                logo_img.add_header("Content-ID", "<company_logo>")
                logo_img.add_header(
                    "Content-Disposition", "inline", filename="company-logo.png"
                )
                email.mixed_subtype = "related"
                email.attach(logo_img)
            except Exception:  # noqa: BLE001 — a logo failure must not block the mail
                logger.warning("Could not attach inline logo from %s", logo_path)

        email.attach(
            filename,
            xlsx_bytes,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        _send_email_with_retry(email, report_date)

        logger.info(
            "Daily production report email sent for %s to=%d cc=%d",
            report_date, len(to_list), len(cc_list),
        )
        return True

    except Exception:  # noqa: BLE001 — mail must never crash the caller
        logger.exception(
            "Failed to send daily production report email for %s", report_date
        )
        return False


# ---------------------------------------------------------------------------
# Live (re)scheduling — called on startup and whenever the admin saves the
# ReportScheduleConfig row, so the send time changes without a server restart.
# ---------------------------------------------------------------------------
DAILY_PRODUCTION_JOB_ID = "daily_production_report_email"
_DEFAULT_HOUR = 18
_DEFAULT_MINUTE = 0


def reschedule_daily_production_job(scheduler) -> None:
    """
    Remove the existing daily-production job and re-add it from the current
    ReportScheduleConfig row. If the config is disabled, the job is left removed.
    Falls back to 18:00 when the config row/table isn't available yet (e.g. on
    first startup before migrations run). Never raises.
    """
    try:
        from apscheduler.triggers.cron import CronTrigger

        # Drop any existing job first so a changed time / disable takes effect.
        if scheduler.get_job(DAILY_PRODUCTION_JOB_ID):
            scheduler.remove_job(DAILY_PRODUCTION_JOB_ID)

        hour, minute, enabled = _DEFAULT_HOUR, _DEFAULT_MINUTE, True
        try:
            from tracking.models.mail import ReportScheduleConfig, ReportType

            config = ReportScheduleConfig.objects.filter(
                report_type=ReportType.DAILY_PRODUCTION
            ).first()
            if config is not None:
                hour, minute, enabled = (
                    config.send_hour,
                    config.send_minute,
                    config.is_enabled,
                )
        except Exception:  # noqa: BLE001 — table may not exist pre-migration
            logger.warning(
                "ReportScheduleConfig unavailable; using default %02d:%02d.",
                _DEFAULT_HOUR, _DEFAULT_MINUTE,
            )

        if not enabled:
            logger.info(
                "Daily production report schedule is disabled; job removed."
            )
            return

        # Import the job entry point lazily to avoid an import cycle with apps.py.
        from tracking.apps import _run_daily_production_report_job

        scheduler.add_job(
            _run_daily_production_report_job,
            trigger=CronTrigger(hour=hour, minute=minute),
            id=DAILY_PRODUCTION_JOB_ID,
            replace_existing=True,
            max_instances=1,
            coalesce=True,
        )
        logger.info(
            "Daily production report job scheduled at %02d:%02d Asia/Dhaka.",
            hour, minute,
        )
    except Exception:  # noqa: BLE001 — scheduling must never crash the caller
        logger.exception("Failed to (re)schedule the daily production report job.")
